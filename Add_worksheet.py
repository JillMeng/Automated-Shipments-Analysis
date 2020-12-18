import pandas as pd
from openpyxl import load_workbook

def add_worksheet(df,filename,sheetname):
    '''
    load workbook and create a new workbook
    :return: new workbook
    '''
    wb = load_workbook('{}.xlsx'.format(filename))
    writer = pd.ExcelWriter('{}.xlsx'.format(filename), engine='openpyxl')
    df.to_excel(wb, sheet_name=sheetname, startrow=1, header=False, index=False)

    worksheet = wb.sheets[sheetname]
    # Get the dimensions of the dataframe.
    (max_row, max_col) = df.shape
    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df.columns]
    # Add the Excel table structure. Pandas will add the data.
    worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

    # Make the columns wider for clarity.
    worksheet.set_column(0, max_col - 1, 12)
    writer.save()
    wb.save()

    '''
    import pandas
    from openpyxl import load_workbook

    book = load_workbook('Masterfile.xlsx')
    writer = pandas.ExcelWriter('Masterfile.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

    writer.save()
    '''